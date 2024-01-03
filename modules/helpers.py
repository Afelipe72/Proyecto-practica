import shutil

import numpy as np
import math
import supervision as sv
from supervision import Point
from ultralytics import YOLO
import csv
import openpyxl
from openpyxl import load_workbook
import cv2
from functools import partial
# from modules.GUI import GSD_user_input

from modules.variables import *

from modules.constants import *
import modules.variables
from modules.variables import *
from modules.variables import current_frame
from modules.variables import format_time_elapsed
from modules.files import *
from modules.files import excel_file_path, path_report_copy

def get_center_bounding_box(coordinates_in_detection) -> list:
    x_center = (coordinates_in_detection[0] + coordinates_in_detection[2]) / 2
    y_center = (coordinates_in_detection[1] + coordinates_in_detection[3]) / 2
    bounding_box_center = [x_center, y_center]
    return bounding_box_center


def calculate_speed(previous_frame_coordinates, current_frame_coordinates) -> dict:
    c_speeds = {}
    for key in current_frame_coordinates:
        if key in previous_frame_coordinates:
            prev_x, prev_y = previous_frame_coordinates[key]
            current_x, current_y = current_frame_coordinates[key]
            x_change = (current_x - prev_x) * GSD / frames_per_second
            y_change = (current_y - prev_y) * GSD / frames_per_second
            speed_vehicle = math.sqrt(x_change ** 2 + y_change ** 2)
            c_speeds[key] = (x_change, y_change, speed_vehicle)
        else:
            c_speeds[key] = (0, 0, 0)
    return c_speeds


def calculate_acceleration(speed_x_y_total_previous_frame, speed_x_y_total_current_frame) -> dict:
    acceleration = {}
    for key in speed_x_y_total_current_frame:
        if key in speed_x_y_total_previous_frame:
            x_value_previous, y_value_previous, total_val_previous = speed_x_y_total_previous_frame[key]
            x_value_current, y_value_current, total_val_current = speed_x_y_total_current_frame[key]
            x_change = ((x_value_current - x_value_previous) / frames_per_second)*0.1
            y_change = ((y_value_current - y_value_previous) / frames_per_second)*0.1
            acceleration_vehicle = math.sqrt(x_change ** 2 + y_change ** 2)
            acceleration[key] = (x_change, y_change, acceleration_vehicle)
        else:
            acceleration[key] = (0, 0, 0)
    return acceleration


def format_csv_values(tracker_id, bounding_box_anchors, class_id, confidence_value) -> list:
    global coordinates, previous_frame_bounding_box, previous_speeds_to_acceleration, GSD, \
         vx, vy, vt, ax, ay, at
    car_id = tracker_id
    bounding_box_center = get_center_bounding_box(bounding_box_anchors)
    coordinates[car_id] = bounding_box_center

    speed_to_csv = calculate_speed(previous_frame_bounding_box, coordinates)
    accelerations_to_csv = calculate_acceleration(previous_speeds_to_acceleration, speed_to_csv)

    for _ in speed_to_csv:
        if car_id in speed_to_csv:
            vx, vy, vt = speed_to_csv[car_id]

    for _ in accelerations_to_csv:
        if car_id in accelerations_to_csv:
            ax, ay, at = accelerations_to_csv[car_id]

    csv_values = [current_frame, format_time_elapsed, class_id,
                  confidence_value * 100, car_id, bounding_box_center[0],
                  bounding_box_center[1], bounding_box_center[0] * GSD, bounding_box_center[1] * GSD,
                  vx, vy, vt, ax, ay, at
                  ]
    # speed
    for car_id in coordinates:
        if car_id not in previous_frame_bounding_box:
            previous_frame_bounding_box[car_id] = coordinates[car_id]

    for car_id in coordinates:
        previous_frame_bounding_box[car_id] = coordinates[car_id]

    # acceleration
    for car_id in speed_to_csv:
        if car_id not in previous_speeds_to_acceleration:
            previous_speeds_to_acceleration[car_id] = speed_to_csv[car_id]

    for car_id in speed_to_csv:
        previous_speeds_to_acceleration[car_id] = speed_to_csv[car_id]

    return csv_values


def write_values_on_csv_raw(csv_values):
    global header_written_raw_csv
    if not header_written_raw_csv:
        vehicles_csv_header = ['Frame', 'Tiempo(S)', 'Tipo de vehiculo', '% Certeza', 'Id Vehículo', 'X (Coordenadas)',
                               'Y (Coordenadas)', 'X (m)', 'Y (m)', 'Vx (m/s)', 'Vy(m/s)', 'Vt(m/s)', 'Ax(m/s)',
                               'Ay(m/s)', 'At(m/s)']
        with open('raw.csv', 'a', newline='') as vehicles_csv:
            writer = csv.writer(vehicles_csv)
            writer.writerow(vehicles_csv_header)

        header_written_raw_csv = True

    # Open the CSV file in append mode and write the values
    with open('raw.csv', 'a', newline='') as vehicles_csv:
        writer = csv.writer(vehicles_csv)
        writer.writerow(csv_values)


def polygon_zone() -> dict:
    polygon_zone_dict = {}
    video_info = sv.VideoInfo.from_video_path("vehicle-counting.mp4")
    number_of_entries = int(input("number of entries"))

    zones_dict = {}  # Dictionary to store zones and their annotators

    for i in range(1, number_of_entries + 1):
        polygon_zone_name = input('polygon zone name\n')
        x1 = int(input('x1\n '))
        y1 = int(input('y1\n '))

        x2 = int(input('x2\n '))
        y2 = int(input('y2\n '))

        x3 = int(input('x3\n '))
        y3 = int(input('y3\n '))

        x4 = int(input('x4\n '))
        y4 = int(input('y4\n '))
        polygon_zone_dict[polygon_zone_name] = np.array([[x1, y1], [x2, y2], [x3, y3], [x4, y4]])

        zone = sv.PolygonZone(polygon=polygon_zone_dict[polygon_zone_name], frame_resolution_wh=video_info.resolution_wh)
        zone_annotator = sv.PolygonZoneAnnotator(zone=zone, color=sv.Color(0, 0, 255), thickness=2, text_thickness=2, text_scale=2)

        zones_dict[polygon_zone_name] = {
            'zone': zone,
            'annotator': zone_annotator,
            'tracked_vehicles': set(),
            'vehicle_count': 0,
            'polygon_zone_dict': [x1, y1]
        }

        create_excel_sheets(zones_dict)

    return zones_dict


def create_excel_sheets(number_of_zones):
    shutil.copy(modules.files.excel_file_path, modules.files.path_report_copy)
    wb = load_workbook(modules.files.path_report_copy)
    target_worksheet = wb['template']
    for key, zone_items in number_of_zones.items():
        new_sheet = wb.copy_worksheet(target_worksheet)
        new_sheet.title = key
    del wb['template']
    wb.save(modules.files.path_report_copy)


zones_dict = polygon_zone()
zone_vehicle_count = {zone_name: 0 for zone_name in zones_dict.keys()}


def process_polygon_zone(zones_dict, detections, frame):
    for zone_name, zone_data in zones_dict.items():
        zone = zone_data['zone']
        zone_annotator = zone_data['annotator']
        tracked_vehicles = zone_data['tracked_vehicles']
        zone_coordinates = zone_data['polygon_zone_dict']
        vehicle_count = zone_vehicle_count[zone_name]

        zone.trigger(detections=detections)
        mask = zone.trigger(detections=detections)
        in_zone_detections = detections[mask]

        zone_objects = modules.variables.processed_objects.get(zone_name, [])

        for i, detection in enumerate(in_zone_detections):
            tracker_id = detection[4]

            if tracker_id not in tracked_vehicles:
                # Increment the vehicle count
                zone_vehicle_count[zone_name] = 1
                tracked_vehicles.add(tracker_id)

                class_id = int(detection[3])
                class_names = CLASS_NAMES_DICT.get(class_id)
                confidence = round(float(detection[2]) * 100)

                data_to_accumulate = {
                    'vehicle_count': zone_vehicle_count[zone_name],
                    'tracker_id': tracker_id,
                    'class_id': class_id,
                    'class_names': class_names,
                    'confidence': confidence
                }

                zone_objects.append(data_to_accumulate)
                # Store the objects for the current zone in the dictionary
            modules.variables.processed_objects[zone_name] = zone_objects

        cv2.putText(frame, zone_name, zone_coordinates, cv2.FONT_HERSHEY_SIMPLEX, 2, (255, 0, 0), 2, cv2.LINE_AA)
        frame = zone_annotator.annotate(scene=frame)

    return modules.variables.processed_objects


def write_csv_polygon_zone(accumulated_data):
    global header_written_zone
    car_values_list = []

    car_value_list_report = []
    if not header_written_zone:
        # minuto , topologia , tipo carro, nombre de la zona, conteo
        vehicles_csv_header = ['Zona', 'Minuto', 'Conteo general', 'Tipología del vehiculo']
        with open('csv_routes.csv', 'a', newline='') as polygon_zone_csv:
            writer = csv.writer(polygon_zone_csv)
            writer.writerow(vehicles_csv_header)

        header_written_zone = True

    for key, car_values in accumulated_data.items():
        for value in car_values:
            # items_accumulated_data = {'zone':key, **value}
            zone_name = key
            vehicle_count = value['vehicle_count']
            tracker_id = value['tracker_id']
            class_id = value['class_id']
            class_names = value['class_names']
            confidence = value['confidence']

            # Append the extracted values to the list
            car_values_list.extend([zone_name, modules.variables.format_time_elapsed, vehicle_count, class_names])

            car_value_list_report.extend([zone_name, modules.variables.format_time_elapsed, vehicle_count, class_names])

            # Open the CSV file in append mode and write the values
            with open('csv_routes.csv', 'a', newline='') as polygon_zone_csv:
                writer = csv.writer(polygon_zone_csv)
                writer.writerow(car_values_list)

            # Clear the list for the next row
            car_values_list = []

    format_report_values(car_value_list_report)


def format_report_values(car_zone_value_list):
    # Assuming the vehicle types are always at every 4th position in the list
    vehicle_types = car_zone_value_list[3::4]
    # Assuming the zone timers are always at every 4th position in the list
    zone_timers = car_zone_value_list[1::4]
    # Assuming the zone names are always at every 4th position in the list
    zone_names = car_zone_value_list[::4]


    # Create a dictionary to store counts and zone timers for each zone
    zone_counts = {}

    for vehicle_type, zone_timer, zone_name in zip(vehicle_types, zone_timers, zone_names):
        if zone_name not in zone_counts:
            # If the zone name is not in the dictionary, add it
            zone_counts[zone_name] = {}

        if vehicle_type not in zone_counts[zone_name]:
            # If the vehicle type is not in the inner dictionary, add it
            zone_counts[zone_name][vehicle_type] = {'Zone timer': zone_timer, 'Count': 0}

        # Add 1 to the count for the corresponding vehicle type in the given zone
        zone_counts[zone_name][vehicle_type]['Count'] += 1

    # Pass the argument to the write_excel_file method
    write_excel_file(zone_counts)



counter_header_file = 0
header_written_excel_file = False
counter_zone_timer = 0
vehicle_counter_zone_timer = 0


def write_excel_file(vehicle_type_counts):
    global excel_file_path, wb, ws, vehicles_saved_in_write_excel, save_vehicles_for_header, CLASS_NAMES_DICT, \
        counter_header_file, header_written_excel_file, counter_zone_timer, vehicle_counter_zone_timer

    print(vehicle_type_counts)
    # Write the vehicles
    # {'test': {'carros': {'Zone timer': ' 1.000', 'Count': 8}},
    # 'test2': {'carros': {'Zone timer': ' 1.000', 'Count': 1}}}

    # for zone_name, zone_items in vehicle_type_counts.items():
    #     print(f"Zone_name: {zone_name}")
    #     for vehicle_type, zone_timer_and_count in zone_items.items():
    #         print(f"Vehicle Type: {vehicle_type}")
    #         for zone_timer, vehicle_count in zone_timer_and_count.items():
    #             print(f"Zone Timer: {zone_timer}")
    #             print(f"Count: {vehicle_count}")


    # Write the header for the file
    for zone_name in vehicle_type_counts.items():
            if not header_written_excel_file:
                start_column_index = 3
                # start column index to set the start value
                for col_idx, (key, value) in enumerate(CLASS_NAMES_DICT.items(), start=start_column_index):
                    # Col idx to increase the column
                    for row in ws.iter_rows(min_row=9, min_col=col_idx, max_row=9, max_col=col_idx):
                        for cell in row:
                            cell.value = value
                wb.save('FORMATO PARA AFORO DE VEHICULOS.xlsx')
                header_written_excel_file = True


    # Time zone for the Excel file
    timer = 0
    for zone_name, zone_items in vehicle_type_counts.items():
        for vehicle_type, zone_timer_and_count in zone_items.items():
            timer = zone_timer_and_count['Zone timer']
    for row in ws.iter_rows(min_row=counter_zone_timer + 10, min_col=2, max_row=counter_zone_timer+10, max_col=2):
        for cell in row:
            cell.value = timer
        wb.save('FORMATO PARA AFORO DE VEHICULOS.xlsx')
        counter_zone_timer += 1



    # Vehicle writer counter
    start_column_index_counter = 3
    vehicle_counter_tracker = 0
    # Puts the vehicle type in a list
    header_order = [CLASS_NAMES_DICT[i] for i in range(len(CLASS_NAMES_DICT))]

    # start column index to set the start value
    for zone_name, zone_items in vehicle_type_counts.items():
        for vehicle_type, zone_timer_and_count in zone_items.items():
            # Find the index of the current vehicle type in the header order
            col_idx = header_order.index(vehicle_type) + start_column_index_counter

            # Get the count value for the current vehicle type
            vehicle_counter_tracker = zone_timer_and_count['Count']

            # Write the count value to the corresponding cell in the Excel file
            for row_idx in range(vehicle_counter_zone_timer + 10, vehicle_counter_zone_timer + 11):
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=row_idx, max_row=row_idx):
                    for c in cell:
                        c.value = vehicle_counter_tracker

            wb.save('FORMATO PARA AFORO DE VEHICULOS.xlsx')
    vehicle_counter_zone_timer += 1





tracker = sv.ByteTrack()
box_annotator = sv.BoundingBoxAnnotator()
label_annotator = sv.LabelAnnotator()
trace_annotator = sv.TraceAnnotator()
