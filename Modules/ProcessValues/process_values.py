import supervision as sv
import numpy as np
import cv2
import openpyxl
from openpyxl import load_workbook
# Method: process_polygon_zone
from Modules.Values.variables import processed_objects
# from Modules.Values.declare_polygon_zone import zone_vehicle_count

from ultralytics import YOLO


# Method: callback
import Modules
from Modules.Values.files import model
from Modules.Values.variables import current_frame, time_elapsed_reset, time_elapsed, zone_timer  # processed_objects

#from Modules.Values.declare_polygon_zone import zones_dict
from Modules.FormatValues.format_values import format_csv_values
from Modules.WriteValuesOnFiles.write_values_on_files import write_values_on_csv_raw, write_csv_polygon_zone
from Modules.Values.constants import CLASS_NAMES_DICT

# Method: process_polygon_zone
from Modules.Values.variables import processed_objects
#from Modules.Values.declare_polygon_zone import zone_vehicle_count
from Modules.Values.constants import CLASS_NAMES_DICT
from Modules.Values.files import excel_file_path_coordinates


# Method: polygon_zone
from Modules.WriteValuesOnFiles.write_values_on_files import create_excel_sheets
#
fps = 0

box_annotator = None
label_annotator = None
round_box_annotator = None
trace_annotator = None
corner_annotator = None
color_annotator = None
circle_annotator = None
dot_annotator = None
triangle_annotator = None
ellipse_annotator = None
percentage_bar_annotator = None
blur_annotator = None
pixelate_annotator = None
heat_map_annotator = None
selected_annotators = []

def gooey_receiver(args=None):
    global fps, box_annotator, label_annotator, round_box_annotator, trace_annotator, corner_annotator, color_annotator, \
        dot_annotator, triangle_annotator,circle_annotator, ellipse_annotator, percentage_bar_annotator, blur_annotator, pixelate_annotator, heat_map_annotator, selected_annotators
#     # Receives the Excel file with the coordinates
    Modules.Values.files.excel_file_path_coordinates = f"{args.Coordenadas}"
    # Receives the Excel file template
    Modules.Values.files.excel_file_path = f"{args.Plantilla}"
    # Receives the Yolo model
    Modules.Values.files.model = YOLO(f"{args.Modelo}")
    Modules.Values.constants.CLASS_NAMES_DICT = Modules.Values.files.model.model.names
    # Receives the GSD
    Modules.Values.constants.GSD = args.GSD
    # Sets the video resolution
    Modules.Values.files.video_info_resolution = f"{args.Video}"
    # Sets raw time threshold
    Modules.Values.constants.user_input_minutes_raw_csv = args.Frecuencia
    # Sets zone time threshold
    Modules.Values.constants.user_input_minutes_zone_timer = args.Rutas * 60
    # Gets the videos fps
    video_test = cv2.VideoCapture(f"{args.Video}")
    fps = round(video_test.get(cv2.CAP_PROP_FPS))
    Modules.Values.constants.frames_per_second = 1 / fps

    # Custom labels
    if args.BoundingBox:
        box_annotator = sv.BoundingBoxAnnotator()

    if args.Label:
        label_annotator = sv.LabelAnnotator()

    if args.Trace:
        trace_annotator = sv.TraceAnnotator(trace_length=9000)

    if args.BoxCorner:
        corner_annotator = sv.BoxCornerAnnotator()

    if args.Color:
        color_annotator = sv.ColorAnnotator()

    if args.Circle:
        circle_annotator = sv.CircleAnnotator()

    if args.Dot:
        dot_annotator = sv.DotAnnotator()

    if args.Triangle:
        triangle_annotator = sv.TriangleAnnotator()

    if args.Ellipse:
        ellipse_annotator = sv.EllipseAnnotator()

    if args.Blur:
        blur_annotator = sv.BlurAnnotator()

    if args.HeatMap:
        # heat_map_annotator = sv.HeatMapAnnotator(radius=10, kernel_size=25)
        heat_map_annotator = sv.HeatMapAnnotator(radius=10, kernel_size=25)

    selected_annotators = [
        box_annotator, label_annotator, circle_annotator, trace_annotator,
        corner_annotator, color_annotator, dot_annotator, triangle_annotator,
        ellipse_annotator, percentage_bar_annotator, blur_annotator,
        pixelate_annotator, heat_map_annotator
    ]

    # Process the video
    sv.process_video(
        source_path=f"{args.Video}",
        target_path="result.mp4",
        callback=callback
    )


def polygon_zone() -> dict:
    polygon_zone_dict = {}
    zones_dict = {}

    try:
        wb = load_workbook(Modules.Values.files.excel_file_path_coordinates)
        ws = wb.active

        num_rows = ws.max_row
        num_columns = ws.max_column

        video_info = sv.VideoInfo.from_video_path(Modules.Values.files.video_info_resolution)

        for row in ws.iter_rows(min_row=2, min_col=1, max_row=num_rows, max_col=num_columns):
            zone_coordinates = []
            polygon_zone_name = None
            for cell in row:
                if cell.column == 1:
                    polygon_zone_name = cell.value
                else:
                    zone_coordinates.append(int(cell.value))
            if polygon_zone_name is not None:
                # Assuming the values are in pairs (x, y), adjust accordingly
                values = np.array(list(zip(zone_coordinates[::2], zone_coordinates[1::2])))
                polygon_zone_dict[polygon_zone_name] = values

                zone = sv.PolygonZone(polygon=polygon_zone_dict[polygon_zone_name], frame_resolution_wh=video_info.resolution_wh)
                zone_annotator = sv.PolygonZoneAnnotator(zone=zone, color=sv.Color(255, 225, 6), thickness=2, text_thickness=1, text_scale=1, text_padding=0)

                zones_dict[polygon_zone_name] = {
                    'zone': zone,
                    'annotator': zone_annotator,
                    'tracked_vehicles': set(),
                    'vehicle_count': 0,
                    'zone_coordinates_polygon': values,
                    'zone_name_coordinates': [int(zone_coordinates[2]), int(zone_coordinates[3])]
                }

                create_excel_sheets(zones_dict)

    except Exception as e:
        # Handle the case where the Excel file couldn't be loaded
        print("No se detectó un archivo con las coordenadas.")
        return {}

    return zones_dict


tracker_id_to_zone_id = {}
counts = {}

def update_tracker_info(detections_in_zones, detections) -> sv.Detections:
    global tracker_id_to_zone_id, counts
    detections_all = detections
    for zone_id, detections_zone in enumerate(detections_in_zones):
        for tracker_id in detections_zone.tracker_id:
            tracker_id_to_zone_id.setdefault(tracker_id, zone_id)

    for zone_id, detections_zone in enumerate(detections_in_zones):
        for tracker_id in detections_zone.tracker_id:
            if tracker_id in tracker_id_to_zone_id:
                zone_in_id = tracker_id_to_zone_id[tracker_id]
                counts.setdefault(zone_id, {})
                counts[zone_id].setdefault(zone_in_id, set())
                counts[zone_id][zone_in_id].add(tracker_id)

    # Update class_id in detections_all based on tracker_id_to_zone_id
    detections_all.class_id = np.vectorize(
        lambda x: tracker_id_to_zone_id.get(x, -1)
    )(detections_all.tracker_id)

    # Filter out detections with class_id == -1
    return detections_all[detections_all.class_id != -1]



def process_polygon_zone(zones_dict, detections, frame):
    zone_vehicle_count = {zone_name: 0 for zone_name in zones_dict.keys()}
    # Testing traffic analisis
    detections_in_zone = []

    for zone_name, zone_data in zones_dict.items():
        zone = zone_data['zone']
        tracked_vehicles = zone_data['tracked_vehicles']
        zone_name_coordinates = zone_data['zone_name_coordinates']
        zone_coordinates_polygon = zone_data['zone_coordinates_polygon']
        vehicle_count = zone_vehicle_count[zone_name]

        zone.trigger(detections=detections)
        mask = zone.trigger(detections=detections)
        in_zone_detections = detections[mask]
        zone_objects = Modules.Values.variables.processed_objects.get(zone_name, [])

        # Accumulate detections for each zone
        for i, detection in enumerate(in_zone_detections):
            tracker_id = detection[4]

            if tracker_id not in tracked_vehicles:
                # Increment the vehicle count
                zone_vehicle_count[zone_name] = 1
                tracked_vehicles.add(tracker_id)

                class_id = int(detection[3])
                class_names = Modules.Values.constants.CLASS_NAMES_DICT.get(class_id)
                confidence = round(float(detection[2]) * 100)

                data_to_accumulate = {
                    'vehicle_count': zone_vehicle_count[zone_name],
                    'tracker_id': tracker_id,
                    'class_id': class_id,
                    'class_names': class_names,
                    'confidence': confidence
                }

                zone_objects.append(data_to_accumulate)

        # draw in the polygon
        if in_zone_detections:
            # Testing blinking when vehicles are inside the zone
            overlay = frame.copy()
            cv2.fillPoly(overlay, np.int32([zone_coordinates_polygon]), color=(255, 255, 255))
            # Transparency factor
            alpha = 0.2
            cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0, frame)

        # Store the objects for the current zone in the dictionary
        Modules.Values.variables.processed_objects[zone_name] = zone_objects

        # Testing traffic analysis
        # Accumulate detections for all zones
        detections_in_zone.append(in_zone_detections)

        cv2.putText(frame, zone_name, zone_name_coordinates, cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2, cv2.LINE_AA)
        frame = zone_data['annotator'].annotate(scene=frame)

    return Modules.Values.variables.processed_objects, detections_in_zone


tracker = sv.ByteTrack()

zones_dict = None


def callback(frame: np.ndarray, _: int) -> np.ndarray:
    global zones_dict, fps, selected_annotators, label_annotator
    if not Modules.Values.variables.header_written_process_polygon_zone:
        zones_dict = polygon_zone()
        Modules.Values.variables.header_written_process_polygon_zone = True
    # Process frame
    results = Modules.Values.files.model(frame)[0]
    detections = sv.Detections.from_ultralytics(results)
    detections = tracker.update_with_detections(detections)

    Modules.Values.variables.current_frame += 1

    Modules.Values.variables.time_elapsed_reset += 1 / fps
    format_time_elapsed_reset = f"{Modules.Values.variables.time_elapsed_reset: 0.3f}"
    format_time_elapsed_reset_to_float = float(format_time_elapsed_reset)

    # sv.plot_image(image=frame, size=(16, 16))

    # For the raw_csv without resetting
    Modules.Values.variables.time_elapsed += 1 / fps
    Modules.Values.variables.format_time_elapsed = f"{Modules.Values.variables.time_elapsed: 0.3f}"

    # Process for the raw csv Values
    if abs(format_time_elapsed_reset_to_float - Modules.Values.constants.user_input_minutes_raw_csv) <= (1/fps)*5:
        for car_value in range(len(detections)):
            values_to_csv = format_csv_values(detections.tracker_id[car_value], detections.xyxy[car_value],
                            Modules.Values.constants.CLASS_NAMES_DICT.get(detections.class_id[car_value]),detections.confidence[car_value])
            # writes value
            if abs(format_time_elapsed_reset_to_float - Modules.Values.constants.user_input_minutes_raw_csv) < 0.01:  # Modules.Values.constants.user_input_minutes_raw_csv
                write_values_on_csv_raw(values_to_csv)

    # resets value
    print(format_time_elapsed_reset_to_float)
    if abs(format_time_elapsed_reset_to_float - Modules.Values.constants.user_input_minutes_raw_csv) < 0.01: # Modules.Values.constants.user_input_minutes_raw_csv
        Modules.Values.variables.time_elapsed_reset = 0

    # Process the polygon zone
    process_polygon_zone(zones_dict, detections, frame)

    # Polygon zone timer
    Modules.Values.variables.zone_timer += 1 / fps
    Modules.Values.variables.format_time_elapsed_test = f"{Modules.Values.variables.zone_timer: 0.3f}"
    print(Modules.Values.variables.format_time_elapsed_test)
    if abs(float(Modules.Values.variables.format_time_elapsed_test) - Modules.Values.constants.user_input_minutes_zone_timer) < 0.01:
        write_csv_polygon_zone(process_polygon_zone(zones_dict, detections, frame))
        Modules.Values.variables.processed_objects = {}
        Modules.Values.variables.zone_timer = 0

    processed_objects_test, detections_in_zone = process_polygon_zone(zones_dict, detections, frame)
    # decomentar para la detección solo en poligonos
    # detections = update_tracker_info(detections_in_zone, detections)

    labels = [
        f"#{tracker_id} {results.names[class_id]}"
        for class_id, tracker_id
        in zip(detections.class_id, detections.tracker_id)
    ]
    # annotated_frame_label = label_annotator.annotate
    annotated_frame = frame.copy()
    selected_annotators = [annotator for annotator in selected_annotators if annotator is not None]

    cv2.imshow('Proceso del video', annotated_frame)

    for annotator in selected_annotators:
        annotated_frame = annotator.annotate(annotated_frame, detections=detections)
        if label_annotator:
            label_annotator.annotate(annotated_frame, detections=detections, labels=labels)

    # Display the annotated frame using OpenCVv
    img_resize = cv2.resize(annotated_frame.copy(), (780, 780))
    cv2.imshow('Proceso del video', img_resize)

    # Wait for a key press to avoid overwhelming the display
    cv2.waitKey(2)  # Adjust delay as needed (higher value for slower processing)

    return annotated_frame







