from openpyxl import load_workbook


# Method: write_excel_file
import Modules
from Modules.Values.files import path_report_copy
from Modules.Values.constants import CLASS_NAMES_DICT

def write_excel_file(vehicle_type_counts):
    print(vehicle_type_counts)
    # Write the vehicles
    # {'test': {'carros': {'Zone timer': ' 1.000', 'Count': 8}},
    # 'test2': {'carros': {'Zone timer': ' 1.000', 'Count': 1}}}

    # for zone_name, zone_items in vehicle_type_counts.items():
    #     print(f"Zone_name: {zone_name}")
    #     for vehicle_type, zone_timer_and_count in zone_items.items():
    #         print(f"Vehicle Type: {vehicle_type}")
    #         for zone_timer, vehicle_count in zone_timer_and_count.items():
    #             print(f"Zone Timer: {zone_timer}")
    #             print(f"Count: {vehicle_count}")
    wb = load_workbook(path_report_copy)
    ws = wb.active

    # Write the header for the file
    for zone_name, _ in vehicle_type_counts.items():
        for sheet in wb.worksheets:
            if zone_name == sheet.title:
                ws = wb[zone_name]
                start_column_index = 3
                # start column index to set the start value
                for col_idx, (key, value) in enumerate(Modules.Values.constants.CLASS_NAMES_DICT.items(), start=start_column_index):
                    # Col idx to increase the column
                    for row in ws.iter_rows(min_row=9, min_col=col_idx, max_row=9, max_col=col_idx):
                        for cell in row:
                            cell.value = value
    wb.save(path_report_copy)

    # Time zone for the Excel file
    timer = 0
    for zone_name_w, _ in vehicle_type_counts.items():
        for sheet in wb.worksheets:
            if zone_name_w == sheet.title:
                ws = wb[zone_name_w]
                for zone_name, zone_items in vehicle_type_counts.items():
                    for vehicle_type, zone_timer_and_count in zone_items.items():
                        timer = float(zone_timer_and_count['Zone timer'])/60
                for row in ws.iter_rows(min_row=Modules.Values.variables.counter_zone_timer + 10, min_col=2, max_row=Modules.Values.variables.counter_zone_timer+10, max_col=2):
                    for cell in row:
                        cell.value = timer
    Modules.Values.variables.counter_zone_timer += 1
    wb.save(path_report_copy)

    # Vehicle writer counter
    start_column_index_counter = 3
    vehicle_counter_tracker = 0
    # Puts the vehicle type in a list
    header_order = [Modules.Values.constants.CLASS_NAMES_DICT[i] for i in range(len(Modules.Values.constants.CLASS_NAMES_DICT))]
    # start column index to set the start value
    for zone_name, zone_items in vehicle_type_counts.items():
        for sheet in wb.worksheets:
            if zone_name == sheet.title:
                ws = wb[zone_name]
                for vehicle_type, zone_timer_and_count in zone_items.items():
                    # Find the index of the current vehicle type in the header order
                    col_idx = header_order.index(vehicle_type) + start_column_index_counter
                    # Get the count value for the current vehicle type
                    vehicle_counter_tracker = zone_timer_and_count['Count']
                    # Write the count value to the corresponding cell in the Excel file
                    for row_idx in range(Modules.Values.variables.vehicle_counter_zone_timer + 10, Modules.Values.variables.vehicle_counter_zone_timer + 11):
                        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=row_idx, max_row=row_idx):
                            for c in cell:
                                c.value = vehicle_counter_tracker
                    wb.save(path_report_copy)
    Modules.Values.variables.vehicle_counter_zone_timer += 1